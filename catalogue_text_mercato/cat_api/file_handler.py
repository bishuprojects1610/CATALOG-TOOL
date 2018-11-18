import boto3
import re
from openpyxl import load_workbook
from .models import Product, Project, User
from .serializer import  ProductSerializer, InputHeadersSerializer, DataFilesSerializer, OutputHeadersSerializer
from threading import Thread


pattern = re.compile('([0-9])')

def s3_upload(_file, dest):
    s3 = boto3.resource('s3')
    upload_file = s3.Bucket('text-mercato-new').put_object(Key=_file.name, Body=_file)
    return True

def document_upload(data_dict, data_type):
    if data_type == 'input_files':
        serializer = ProductSerializer(data=data_dict)
    elif data_type == 'out_put_template':
        serializer = OutputHeadersSerializer(data=data_dict)
    elif data_type == 'data_files':
        serializer = DataFilesSerializer(data=data_dict)
    if serializer.is_valid():
        serializer.save()
    else:
        print (serializer.errors)

def header_iterator(data, row_index, file_name, sheet_name, project_id, header_id, data_type):
    row_dict = {}
    for i in range(1, data.max_column+1):
        row_dict[re.sub(pattern, '', data.cell(row=row_index, column=i).coordinate)] = data.cell(row=row_index, column=i).value
    data_dict = {}
    data_dict['details'] = row_dict
    data_dict['project'] = project_id
    data_dict['file_name'] = file_name
    data_dict['sheet_name'] = sheet_name
    if data_type == 'input_files':
        data_dict['input_headers'] = header_id
    document_upload(data_dict, data_type)
    return data_dict


def read_file(_files, file_type, project_id):
    for _file in _files:
        # print (_file, type(_file), dir(_file))
        wb = load_workbook(_file)
        sheets = wb.sheetnames
        if file_type == 'input_files':
            for sheetname in sheets:
                if sheetname != '__hidden__':
                    sheet = wb[sheetname]
                    headers = {re.sub(pattern, '', sheet.cell(row=1, column=i).coordinate): sheet.cell(row=1, column=i).value for i in range(1,sheet.max_column+1)}
                    header_dict = {'headers': headers, 'file_name': _file.name, 'sheet_name': sheetname, 'project': project_id}
                    header_serializer = InputHeadersSerializer(data=header_dict)
                    if header_serializer.is_valid():
                        header_doc = header_serializer.save()
                    for i in range(2, sheet.max_row+1):
                        # product_details = header_iterator(sheet, i, _file.name, sheetname, project_id, header_doc.id, file_type)
                        t = Thread(target=header_iterator, args=[sheet, i, _file.name, sheetname, project_id, header_doc.id, file_type] )
                        t.start()
                        t.join()
                        
            
                            
            
        
            
        elif file_type == 'data_files':
            pass
        elif file_type == 'out_put_template':
            pass
        # s3_upload(_file, 'test')
    return True

                
    

# def s3_upload(request_name, temp_file, s3):
#     # print ('yes', temp_file[0], temp_file[1].name, dir(temp_file[0]))
#     name_list = []
#     for _ in temp_file:
#         wb = load_workbook(_)
#         print (wb.sheetnames, '*********************')
#         # s3.Bucket('text-mercato-new').put_object(Key=_.name, Body=_)
#         name_list.append('text-mercato-new' + _.name)
#     return True

# def read_file(name, templates, s3):
#     for template in templates: